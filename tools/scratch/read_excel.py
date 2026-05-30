import openpyxl
import os

file_path = r"c:\Users\a9999\Dropbox\CodeBackups\CODING\tools\REF\781估價單_V3.1 111111.xlsx"
out_path = r"c:\Users\a9999\Dropbox\CodeBackups\CODING\tools\scratch\781_extracted.md"

wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
sheet = wb[wb.sheetnames[0]] # 讀取第一個工作表

with open(out_path, "w", encoding="utf-8") as f:
    f.write(f"# 781估價單 內容彙整 ({sheet.title})\n\n")
    f.write("| 行號 | 項目 | 單位 | 數量 | 單價 | 備註 | 工法/對價 |\n")
    f.write("| --- | --- | --- | --- | --- | --- | --- |\n")
    
    for r in range(1, 400):
        row_values = [sheet.cell(row=r, column=c).value for c in range(1, 10)]
        if any(v is not None for v in row_values):
            cleaned = [str(v).replace("\n", " ").strip() if v is not None else "" for v in row_values]
            # 補足 7 個欄位
            while len(cleaned) < 7:
                cleaned.append("")
            f.write(f"| {r} | {cleaned[1]} | {cleaned[2]} | {cleaned[3]} | {cleaned[4]} | {cleaned[5]} | {cleaned[7]} |\n")

wb.close()
print(f"匯出完成，檔案已儲存至: {out_path}")
